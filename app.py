from flask import Flask, request, jsonify, send_file, session, redirect, url_for, make_response
import sqlite3, os, io, openpyxl, hashlib, secrets
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import defaultdict
from functools import wraps

from flask import Flask
import sqlite3, hashlib

app = Flask(__name__)

# ── BASE DE DATOS ─────────────────────
DB_PATH = "database.db"

def get_db():
    conn = sqlite3.connect(DB_PATH)
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()

    c.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now'))
        )
    ''')

    conn.commit()
    conn.close()

# ── RUTA PRINCIPAL ─────────────────────
@app.route("/")
def home():
    return "Hola, estoy online 🚀"

# ── INICIO ─────────────────────────────
if __name__ == "__main__":
    init_db()
    app.run(host="0.0.0.0", port=10000)

# ── BD ──────────────────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def init_db():
    conn = get_db()
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS users (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        username   TEXT UNIQUE NOT NULL,
        email      TEXT UNIQUE NOT NULL,
        password   TEXT NOT NULL,
        created_at TEXT DEFAULT (datetime('now'))
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS transactions (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id     INTEGER NOT NULL,
        type        TEXT NOT NULL CHECK(type IN ('income','expense')),
        amount      REAL NOT NULL,
        category    TEXT NOT NULL,
        description TEXT,
        date        TEXT NOT NULL,
        created_at  TEXT DEFAULT (datetime('now')),
        FOREIGN KEY (user_id) REFERENCES users(id)
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS categories (
        id      INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        name    TEXT NOT NULL,
        type    TEXT NOT NULL CHECK(type IN ('income','expense','both')),
        icon    TEXT DEFAULT '📁',
        color   TEXT DEFAULT '#6366f1',
        UNIQUE(user_id, name),
        FOREIGN KEY (user_id) REFERENCES users(id)
    )''')
    conn.commit()
    conn.close()

DEFAULT_CATEGORIES = [
    ('Salario','income','💼','#10b981'),
    ('Inversiones','income','📈','#8b5cf6'),
    ('Otros ingresos','income','💰','#f59e0b'),
    ('Alimentación','expense','🍔','#ef4444'),
    ('Transporte','expense','🚗','#f97316'),
    ('Vivienda','expense','🏠','#6366f1'),
    ('Salud','expense','⚕️','#ec4899'),
    ('Educación','expense','📚','#14b8a6'),
    ('Entretenimiento','expense','🎬','#a855f7'),
    ('Ropa','expense','👕','#f43f5e'),
    ('Servicios','expense','⚡','#0ea5e9'),
    ('Viajes','expense','✈️','#f59e0b'),
    ('Otros gastos','expense','📦','#64748b'),
]

def seed_user_categories(user_id):
    conn = get_db()
    c = conn.cursor()
    for cat in DEFAULT_CATEGORIES:
        c.execute('INSERT OR IGNORE INTO categories (user_id,name,type,icon,color) VALUES (?,?,?,?,?)',
                  (user_id,) + cat)
    conn.commit()
    conn.close()

# ── AUTH DECORATOR ───────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            if request.path.startswith('/api/'):
                return jsonify({'error': 'No autenticado'}), 401
            return redirect('/login')
        return f(*args, **kwargs)
    return decorated

def current_user_id():
    return session.get('user_id')

# ── HTML TEMPLATE ────────────────────────────────────────────────────────────
HTML = '''<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Finanzas – Control de Gastos</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Sans:ital,wght@0,300;0,400;0,500;1,400&display=swap" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
:root {
  --bg:#0a0e1a; --bg2:#111827; --bg3:#1e2740;
  --border:#1e293b; --border2:#334155;
  --text:#f1f5f9; --text2:#94a3b8; --text3:#64748b;
  --accent:#6ee7b7; --accent2:#34d399;
  --income:#10b981; --expense:#f43f5e;
  --blue:#3b82f6; --purple:#8b5cf6; --gold:#f59e0b;
  --sidebar-w:240px; --radius:14px;
  --shadow:0 4px 24px rgba(0,0,0,.45);
}
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
body{background:var(--bg);color:var(--text);font-family:'DM Sans',sans-serif;min-height:100vh;overflow-x:hidden}
button{cursor:pointer;border:none;background:none;font-family:inherit}
input,select{font-family:inherit}
/* AUTH PAGES */
.auth-page{min-height:100vh;display:flex;align-items:center;justify-content:center;padding:20px;background:var(--bg)}
.auth-page::before{content:'';position:fixed;inset:0;background:radial-gradient(ellipse at 20% 50%,rgba(110,231,183,.06) 0%,transparent 60%),radial-gradient(ellipse at 80% 20%,rgba(139,92,246,.06) 0%,transparent 50%);pointer-events:none}
.auth-box{background:var(--bg2);border:1px solid var(--border2);border-radius:24px;padding:44px 40px;width:min(440px,100%);position:relative;animation:authIn .4s cubic-bezier(.34,1.56,.64,1)}
@keyframes authIn{from{opacity:0;transform:translateY(20px) scale(.97)}to{opacity:1;transform:none}}
.auth-logo{display:flex;align-items:center;gap:10px;margin-bottom:32px;justify-content:center}
.auth-logo-icon{font-size:28px;color:var(--accent)}
.auth-logo-text{font-family:'Syne',sans-serif;font-weight:800;font-size:22px;letter-spacing:.03em}
.auth-title{font-family:'Syne',sans-serif;font-size:24px;font-weight:800;margin-bottom:6px;text-align:center}
.auth-sub{font-size:14px;color:var(--text3);text-align:center;margin-bottom:28px}
.auth-form{display:flex;flex-direction:column;gap:14px}
.auth-group{display:flex;flex-direction:column;gap:6px}
.auth-group label{font-size:11px;font-weight:700;color:var(--text3);text-transform:uppercase;letter-spacing:.08em}
.auth-group input{background:var(--bg3);border:1px solid var(--border2);color:var(--text);padding:12px 16px;border-radius:12px;font-size:14px;outline:none;transition:border-color .15s}
.auth-group input:focus{border-color:var(--accent)}
.auth-btn{background:linear-gradient(135deg,var(--accent),#059669);color:#064e3b;font-weight:800;font-size:15px;padding:14px;border-radius:12px;margin-top:8px;transition:all .2s;width:100%}
.auth-btn:hover{filter:brightness(1.08);transform:translateY(-1px);box-shadow:0 6px 20px rgba(16,185,129,.3)}
.auth-link{text-align:center;font-size:13px;color:var(--text3);margin-top:20px}
.auth-link a{color:var(--accent);font-weight:600;text-decoration:none}
.auth-link a:hover{text-decoration:underline}
.auth-error{background:rgba(244,63,94,.12);border:1px solid rgba(244,63,94,.3);color:#f87171;padding:10px 14px;border-radius:10px;font-size:13px;text-align:center}
.auth-divider{display:flex;align-items:center;gap:10px;margin:4px 0}
.auth-divider::before,.auth-divider::after{content:'';flex:1;height:1px;background:var(--border)}
.auth-divider span{font-size:11px;color:var(--text3)}
/* SIDEBAR */
.sidebar{position:fixed;top:0;left:0;height:100vh;width:var(--sidebar-w);background:var(--bg2);border-right:1px solid var(--border);display:flex;flex-direction:column;z-index:200;transition:transform .3s cubic-bezier(.4,0,.2,1)}
.sidebar-logo{display:flex;align-items:center;gap:10px;padding:26px 22px 22px;border-bottom:1px solid var(--border)}
.logo-icon{font-size:24px;color:var(--accent);line-height:1}
.logo-text{font-family:'Syne',sans-serif;font-weight:800;font-size:18px;letter-spacing:.03em;color:var(--text)}
.sidebar-nav{flex:1;padding:16px 10px;display:flex;flex-direction:column;gap:4px}
.nav-btn{display:flex;align-items:center;gap:10px;width:100%;padding:11px 14px;border-radius:10px;font-size:14px;font-weight:500;color:var(--text2);transition:all .18s;text-align:left}
.nav-btn:hover{background:var(--bg3);color:var(--text)}
.nav-btn.active{background:linear-gradient(135deg,#065f46 0%,#064e3b 100%);color:var(--accent);font-weight:600}
.nav-icon{font-size:16px;width:20px;text-align:center}
.sidebar-footer{padding:16px 10px;border-top:1px solid var(--border);display:flex;flex-direction:column;gap:8px}
.export-btn{display:flex;align-items:center;justify-content:center;gap:8px;width:100%;padding:11px;border-radius:10px;background:linear-gradient(135deg,#1d4ed8,#7c3aed);color:#fff;font-size:14px;font-weight:600;transition:opacity .18s}
.export-btn:hover{opacity:.85}
.logout-btn{display:flex;align-items:center;justify-content:center;gap:8px;width:100%;padding:9px;border-radius:10px;background:rgba(244,63,94,.1);color:#f87171;font-size:13px;font-weight:600;border:1px solid rgba(244,63,94,.2);transition:all .18s}
.logout-btn:hover{background:rgba(244,63,94,.2)}
.overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,.6);z-index:190}
.overlay.active{display:block}
/* MAIN */
.main{margin-left:var(--sidebar-w);min-height:100vh;transition:margin .3s}
.topbar{position:sticky;top:0;background:rgba(10,14,26,.85);backdrop-filter:blur(14px);border-bottom:1px solid var(--border);padding:14px 28px;display:flex;align-items:center;gap:16px;z-index:100}
.menu-btn{display:none;font-size:20px;color:var(--text2);padding:4px 8px;border-radius:8px}
.menu-btn:hover{color:var(--text);background:var(--bg3)}
.topbar-title{font-family:'Syne',sans-serif;font-weight:700;font-size:20px;flex:1}
.topbar-actions{display:flex;align-items:center;gap:12px;flex-wrap:wrap}
.topbar-user{font-size:12px;color:var(--text3);background:var(--bg2);border:1px solid var(--border);padding:6px 12px;border-radius:20px}
.topbar-user span{color:var(--accent);font-weight:600}
.add-btn{background:var(--accent);color:#064e3b;font-weight:700;font-size:14px;padding:9px 20px;border-radius:30px;transition:all .18s}
.add-btn:hover{background:var(--accent2);transform:translateY(-1px)}
.month-picker{display:flex;align-items:center;gap:4px;background:var(--bg2);border:1px solid var(--border);border-radius:10px;padding:6px 10px}
.month-picker button{color:var(--text2);font-size:18px;line-height:1;padding:0 4px;border-radius:4px;transition:color .15s}
.month-picker button:hover{color:var(--accent)}
.month-picker span{font-size:13px;font-weight:500;min-width:90px;text-align:center}
.section{display:none;padding:28px;animation:fadeIn .3s ease}
.section.active{display:block}
@keyframes fadeIn{from{opacity:0;transform:translateY(8px)}to{opacity:1;transform:translateY(0)}}
/* CARDS */
.summary-cards{display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:16px;margin-bottom:24px}
.card{background:var(--bg2);border:1px solid var(--border);border-radius:var(--radius);padding:22px 24px;position:relative;overflow:hidden;transition:transform .2s,box-shadow .2s}
.card:hover{transform:translateY(-2px);box-shadow:var(--shadow)}
.card::before{content:'';position:absolute;top:0;left:0;right:0;height:3px}
.card.income::before{background:var(--income)}
.card.expense::before{background:var(--expense)}
.card.balance::before{background:linear-gradient(90deg,var(--blue),var(--purple))}
.card-label{font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.1em;color:var(--text3);margin-bottom:10px}
.card-amount{font-family:'Syne',sans-serif;font-size:28px;font-weight:800;line-height:1;margin-bottom:6px}
.card.income .card-amount{color:var(--income)}
.card.expense .card-amount{color:var(--expense)}
.card.balance .card-amount{color:var(--text)}
.card-sub{font-size:12px;color:var(--text3)}
.card-icon{position:absolute;right:18px;top:50%;transform:translateY(-50%);font-size:36px;opacity:.12}
/* CHARTS */
.charts-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(340px,1fr));gap:20px}
.chart-wide{grid-column:1 / -1}
.chart-card{background:var(--bg2);border:1px solid var(--border);border-radius:var(--radius);padding:20px 22px;transition:box-shadow .2s}
.chart-card:hover{box-shadow:var(--shadow)}
.chart-card-header{display:flex;align-items:center;justify-content:space-between;margin-bottom:16px;font-family:'Syne',sans-serif;font-weight:700;font-size:14px}
.toggle-btns{display:flex;gap:4px}
.tog{font-size:11px;font-weight:600;padding:5px 12px;border-radius:20px;color:var(--text2);border:1px solid var(--border2);transition:all .15s}
.tog.active{background:var(--accent);color:#064e3b;border-color:transparent}
/* FILTERS */
.filter-bar{display:flex;flex-wrap:wrap;gap:10px;margin-bottom:20px}
.filter-input,.filter-select{background:var(--bg2);border:1px solid var(--border2);color:var(--text);padding:9px 14px;border-radius:10px;font-size:13px;outline:none;transition:border-color .15s}
.filter-input:focus,.filter-select:focus{border-color:var(--accent)}
.filter-input{min-width:160px}
.color-input{padding:5px;width:44px;height:40px;cursor:pointer}
option{background:var(--bg2)}
/* TX LIST */
.tx-list{display:flex;flex-direction:column;gap:8px}
.tx-item{display:flex;align-items:center;gap:14px;background:var(--bg2);border:1px solid var(--border);border-radius:12px;padding:14px 18px;transition:all .18s;animation:fadeIn .25s ease}
.tx-item:hover{border-color:var(--border2);transform:translateX(2px)}
.tx-icon{width:42px;height:42px;border-radius:12px;display:flex;align-items:center;justify-content:center;font-size:20px;flex-shrink:0}
.tx-icon.income{background:rgba(16,185,129,.12)}
.tx-icon.expense{background:rgba(244,63,94,.12)}
.tx-info{flex:1;min-width:0}
.tx-cat{font-weight:600;font-size:14px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.tx-desc{font-size:12px;color:var(--text2);margin-top:2px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.tx-date{font-size:12px;color:var(--text3);white-space:nowrap}
.tx-amount{font-family:'Syne',sans-serif;font-size:16px;font-weight:700;white-space:nowrap}
.tx-amount.income{color:var(--income)}
.tx-amount.expense{color:var(--expense)}
.tx-actions{display:flex;gap:6px;flex-shrink:0}
.tx-btn{width:32px;height:32px;border-radius:8px;display:flex;align-items:center;justify-content:center;font-size:14px;transition:all .15s;border:1px solid transparent}
.tx-btn.edit{color:var(--blue);background:rgba(59,130,246,.1)}
.tx-btn.del{color:var(--expense);background:rgba(244,63,94,.1)}
.tx-btn:hover{filter:brightness(1.2);transform:scale(1.1)}
.empty-state{text-align:center;padding:60px 20px;color:var(--text3);font-size:15px}
.empty-state .big{font-size:48px;margin-bottom:12px}
/* CATEGORIES */
.cat-add-form{display:flex;flex-wrap:wrap;gap:10px;background:var(--bg2);border:1px solid var(--border);border-radius:var(--radius);padding:20px;margin-bottom:20px}
.cat-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(180px,1fr));gap:12px}
.cat-item{background:var(--bg2);border:1px solid var(--border);border-radius:12px;padding:16px;display:flex;align-items:center;gap:12px;transition:all .18s;animation:fadeIn .25s}
.cat-item:hover{border-color:var(--border2)}
.cat-item-icon{font-size:24px}
.cat-item-info{flex:1;min-width:0}
.cat-item-name{font-weight:600;font-size:13px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.cat-item-type{font-size:11px;color:var(--text3);margin-top:2px}
.cat-del{color:var(--text3);font-size:16px;opacity:0;transition:all .15s;flex-shrink:0}
.cat-item:hover .cat-del{opacity:1;color:var(--expense)}
/* MODAL */
.modal-bg{display:none;position:fixed;inset:0;background:rgba(0,0,0,.7);backdrop-filter:blur(6px);z-index:500;align-items:center;justify-content:center}
.modal-bg.active{display:flex}
.modal{background:var(--bg2);border:1px solid var(--border2);border-radius:20px;padding:32px 36px;width:min(480px,calc(100vw - 32px));position:relative;animation:modalIn .25s cubic-bezier(.34,1.56,.64,1)}
@keyframes modalIn{from{opacity:0;transform:scale(.92) translateY(12px)}to{opacity:1;transform:scale(1) translateY(0)}}
.modal-close{position:absolute;top:18px;right:18px;width:32px;height:32px;border-radius:8px;background:var(--bg3);color:var(--text2);font-size:14px;transition:all .15s}
.modal-close:hover{background:var(--expense);color:#fff}
.modal-title{font-family:'Syne',sans-serif;font-size:20px;font-weight:800;margin-bottom:20px}
.modal-type-tabs{display:flex;gap:8px;margin-bottom:24px}
.type-tab{flex:1;padding:11px;border-radius:10px;font-size:14px;font-weight:600;border:2px solid var(--border2);color:var(--text2);transition:all .18s}
.type-tab.active[data-type="expense"]{background:rgba(244,63,94,.15);border-color:var(--expense);color:var(--expense)}
.type-tab.active[data-type="income"]{background:rgba(16,185,129,.15);border-color:var(--income);color:var(--income)}
.modal-form{display:flex;flex-direction:column;gap:16px}
.form-group{display:flex;flex-direction:column;gap:6px}
.form-group label{font-size:12px;font-weight:600;color:var(--text3);text-transform:uppercase;letter-spacing:.06em}
.form-group input,.form-group select{background:var(--bg3);border:1px solid var(--border2);color:var(--text);padding:11px 14px;border-radius:10px;font-size:14px;outline:none;transition:border-color .15s}
.form-group input:focus,.form-group select:focus{border-color:var(--accent)}
.submit-btn{background:linear-gradient(135deg,var(--accent),#059669);color:#064e3b;font-weight:800;font-size:15px;padding:13px;border-radius:12px;margin-top:4px;transition:all .2s}
.submit-btn:hover{filter:brightness(1.08);transform:translateY(-1px);box-shadow:0 6px 20px rgba(16,185,129,.3)}
/* TOAST */
.toast{position:fixed;bottom:24px;right:24px;background:var(--bg3);color:var(--text);padding:12px 22px;border-radius:12px;font-size:14px;font-weight:500;box-shadow:var(--shadow);border-left:4px solid var(--accent);transform:translateY(100px);opacity:0;transition:all .35s cubic-bezier(.34,1.56,.64,1);z-index:999}
.toast.show{transform:translateY(0);opacity:1}
.toast.error{border-left-color:var(--expense)}
/* RESPONSIVE */
@media(max-width:768px){
  .sidebar{transform:translateX(-100%)}
  .sidebar.open{transform:translateX(0)}
  .menu-btn{display:flex;align-items:center}
  .main{margin-left:0}
  .section{padding:16px}
  .topbar{padding:12px 16px}
  .topbar-actions{gap:8px}
  .month-picker span{min-width:72px;font-size:12px}
  .charts-grid{grid-template-columns:1fr}
  .summary-cards{grid-template-columns:1fr 1fr}
  .tx-date{display:none}
  .modal{padding:24px 20px}
  .topbar-user{display:none}
}
@media(max-width:420px){
  .summary-cards{grid-template-columns:1fr}
  .topbar-actions .month-picker{display:none}
}
</style>
</head>
<body>

<aside class="sidebar" id="sidebar">
  <div class="sidebar-logo">
    <span class="logo-icon">◈</span>
    <span class="logo-text">Finanzas</span>
  </div>
  <nav class="sidebar-nav">
    <button class="nav-btn active" data-section="dashboard"><span class="nav-icon">⬡</span> Inicio</button>
    <button class="nav-btn" data-section="transactions"><span class="nav-icon">⇄</span> Movimientos</button>
    <button class="nav-btn" data-section="reports"><span class="nav-icon">◎</span> Reportes</button>
    <button class="nav-btn" data-section="categories"><span class="nav-icon">⊞</span> Categorías</button>
  </nav>
  <div class="sidebar-footer">
    <button class="export-btn" id="exportExcel"><span>⬇</span> Exportar Excel</button>
    <button class="logout-btn" onclick="logout()">⎋ Cerrar sesión</button>
  </div>
</aside>

<div class="overlay" id="overlay" onclick="closeSidebar()"></div>

<main class="main" id="main">
  <header class="topbar">
    <button class="menu-btn" id="menuBtn" onclick="toggleSidebar()">☰</button>
    <div class="topbar-title" id="topbarTitle">Inicio</div>
    <div class="topbar-actions">
      <div class="topbar-user">Hola, <span id="usernameDisplay"></span></div>
      <button class="add-btn" id="openAddBtn">+ Nuevo</button>
      <div class="month-picker">
        <button onclick="changeMonth(-1)">‹</button>
        <span id="monthLabel"></span>
        <button onclick="changeMonth(1)">›</button>
      </div>
    </div>
  </header>

  <section class="section active" id="sec-dashboard">
    <div class="summary-cards" id="summaryCards"></div>
    <div class="charts-grid">
      <div class="chart-card">
        <div class="chart-card-header"><span>Evolución mensual</span></div>
        <canvas id="trendChart" height="220"></canvas>
      </div>
      <div class="chart-card">
        <div class="chart-card-header">
          <span>Por categoría</span>
          <div class="toggle-btns">
            <button class="tog active" data-t="expense" onclick="loadPieChart('expense',this)">Gastos</button>
            <button class="tog" data-t="income" onclick="loadPieChart('income',this)">Ingresos</button>
          </div>
        </div>
        <canvas id="pieChart" height="220"></canvas>
      </div>
      <div class="chart-card chart-wide">
        <div class="chart-card-header"><span>Movimientos del mes</span></div>
        <canvas id="dailyChart" height="160"></canvas>
      </div>
    </div>
  </section>

  <section class="section" id="sec-transactions">
    <div class="filter-bar">
      <input class="filter-input" id="searchInput" type="text" placeholder="Buscar…" oninput="loadTransactions()">
      <select class="filter-select" id="filterType" onchange="loadTransactions()">
        <option value="">Todos</option>
        <option value="income">Ingresos</option>
        <option value="expense">Gastos</option>
      </select>
      <select class="filter-select" id="filterCat" onchange="loadTransactions()">
        <option value="">Todas las categorías</option>
      </select>
      <input class="filter-input" id="filterTo"   type="date" onchange="loadTransactions()">
    </div>
    <div class="tx-list" id="txList"></div>
  </section>

  <section class="section" id="sec-reports">
    <div class="charts-grid">
      <div class="chart-card">
        <div class="chart-card-header"><span>Top gastos por categoría</span></div>
        <canvas id="barCatChart" height="260"></canvas>
      </div>
      <div class="chart-card">
        <div class="chart-card-header"><span>Ingresos por categoría</span></div>
        <canvas id="barIncChart" height="260"></canvas>
      </div>
      <div class="chart-card chart-wide">
        <div class="chart-card-header"><span>Balance acumulado</span></div>
        <canvas id="accumChart" height="160"></canvas>
      </div>
    </div>
  </section>

  <section class="section" id="sec-categories">
    <div class="cat-add-form">
      <input class="filter-input" id="catName"  type="text" placeholder="Nombre de categoría">
      <select class="filter-select" id="catType">
        <option value="expense">Gasto</option>
        <option value="income">Ingreso</option>
        <option value="both">Ambos</option>
      </select>
      <input class="filter-input" id="catIcon"  type="text" placeholder="Emoji 📦" maxlength="4">
      <input class="filter-input color-input" id="catColor" type="color" value="#6366f1">
      <button class="add-btn" onclick="addCategory()">Agregar</button>
    </div>
    <div class="cat-grid" id="catGrid"></div>
  </section>
</main>

<div class="modal-bg" id="modalBg">
  <div class="modal">
    <button class="modal-close" onclick="closeModal()">✕</button>
    <h2 class="modal-title" id="modalTitle">Nueva transacción</h2>
    <div class="modal-type-tabs">
      <button class="type-tab active" data-type="expense" onclick="setType('expense',this)">💸 Gasto</button>
      <button class="type-tab" data-type="income" onclick="setType('income',this)">💰 Ingreso</button>
    </div>
    <div class="modal-form">
      <input type="hidden" id="editId">
      <div class="form-group">
        <label>Monto (S/)</label>
        <input type="number" id="fAmount" step="0.01" min="0.01" placeholder="0.00">
      </div>
      <div class="form-group">
        <label>Categoría</label>
        <select id="fCategory"></select>
      </div>
      <div class="form-group">
        <label>Fecha</label>
        <input type="date" id="fDate">
      </div>
      <div class="form-group">
        <label>Descripción (opcional)</label>
        <input type="text" id="fDescription" placeholder="¿En qué gastaste?">
      </div>
      <button class="submit-btn" id="submitBtn" onclick="submitTransaction()">Guardar</button>
    </div>
  </div>
</div>

<div class="toast" id="toast"></div>

<script>
let currentMonth = new Date().toISOString().slice(0,7);
let currentType  = 'expense';
let categories   = [];
let charts       = {};

document.addEventListener('DOMContentLoaded', async () => {
  // load username
  try {
    const me = await fetch('/api/me').then(r=>r.json());
    document.getElementById('usernameDisplay').textContent = me.username;
  } catch(e){}

  updateMonthLabel();
  await loadCategories();
  switchSection('dashboard');
  document.querySelectorAll('.nav-btn').forEach(b => {
    b.addEventListener('click', () => { switchSection(b.dataset.section); closeSidebar(); });
  });
  document.getElementById('openAddBtn').addEventListener('click', openAddModal);
  document.getElementById('exportExcel').addEventListener('click', exportExcel);
});

async function logout() {
  await fetch('/api/logout', {method:'POST'});
  window.location.href = '/login';
}

function switchSection(name) {
  document.querySelectorAll('.section').forEach(s => s.classList.remove('active'));
  document.querySelectorAll('.nav-btn').forEach(b => b.classList.toggle('active', b.dataset.section === name));
  document.getElementById(`sec-${name}`).classList.add('active');
  const titles = { dashboard:'Inicio', transactions:'Movimientos', reports:'Reportes', categories:'Categorías' };
  document.getElementById('topbarTitle').textContent = titles[name];
  if (name === 'dashboard')    loadDashboard();
  if (name === 'transactions') loadTransactions();
  if (name === 'reports')      loadReports();
  if (name === 'categories')   renderCategories();
}

function updateMonthLabel() {
  const [y, m] = currentMonth.split('-');
  const names  = ['Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre'];
  document.getElementById('monthLabel').textContent = `${names[+m-1]} ${y}`;
}
function changeMonth(delta) {
  const d = new Date(currentMonth + '-01');
  d.setMonth(d.getMonth() + delta);
  currentMonth = d.toISOString().slice(0,7);
  updateMonthLabel();
  const active = document.querySelector('.nav-btn.active')?.dataset.section;
  if (active === 'dashboard')    loadDashboard();
  if (active === 'transactions') loadTransactions();
  if (active === 'reports')      loadReports();
}

function toggleSidebar() {
  document.getElementById('sidebar').classList.toggle('open');
  document.getElementById('overlay').classList.toggle('active');
}
function closeSidebar() {
  document.getElementById('sidebar').classList.remove('open');
  document.getElementById('overlay').classList.remove('active');
}

async function loadCategories() {
  const res = await fetch('/api/categories');
  categories = await res.json();
  populateCategorySelects();
}
function populateCategorySelects() {
  const type = currentType;
  const fCat  = document.getElementById('fCategory');
  const fkCat = document.getElementById('filterCat');
  fCat.innerHTML  = '';
  fkCat.innerHTML = '<option value="">Todas las categorías</option>';
  categories.filter(c => c.type === type || c.type === 'both').forEach(c => {
    fCat.appendChild(new Option(`${c.icon} ${c.name}`, c.name));
  });
  categories.forEach(c => {
    fkCat.appendChild(new Option(`${c.icon} ${c.name}`, c.name));
  });
}
function renderCategories() {
  const grid = document.getElementById('catGrid');
  grid.innerHTML = '';
  categories.forEach(c => {
    const div = document.createElement('div');
    div.className = 'cat-item';
    div.innerHTML = `
      <span class="cat-item-icon">${c.icon}</span>
      <div class="cat-item-info">
        <div class="cat-item-name">${c.name}</div>
        <div class="cat-item-type">${c.type==='income'?'Ingreso':c.type==='expense'?'Gasto':'Ambos'}</div>
      </div>
      <button class="cat-del" onclick="deleteCategory(${c.id})" title="Eliminar">✕</button>`;
    grid.appendChild(div);
  });
}
async function addCategory() {
  const name  = document.getElementById('catName').value.trim();
  const type  = document.getElementById('catType').value;
  const icon  = document.getElementById('catIcon').value.trim() || '📁';
  const color = document.getElementById('catColor').value;
  if (!name) { showToast('Ingresa un nombre de categoría','error'); return; }
  const res = await fetch('/api/categories', {
    method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({name, type, icon, color})
  });
  if (res.ok) {
    document.getElementById('catName').value = '';
    document.getElementById('catIcon').value = '';
    await loadCategories(); renderCategories();
    showToast('Categoría agregada ✓');
  } else {
    const d = await res.json();
    showToast(d.error || 'Error al agregar','error');
  }
}
async function deleteCategory(id) {
  if (!confirm('¿Eliminar esta categoría?')) return;
  await fetch(`/api/categories/${id}`, {method:'DELETE'});
  await loadCategories(); renderCategories();
  showToast('Categoría eliminada');
}

async function loadDashboard() {
  const [summary, trend, daily] = await Promise.all([
    fetch(`/api/stats/summary?month=${currentMonth}`).then(r=>r.json()),
    fetch('/api/stats/monthly_trend').then(r=>r.json()),
    fetch(`/api/stats/daily_trend?month=${currentMonth}`).then(r=>r.json()),
  ]);
  renderSummaryCards(summary);
  renderTrendChart(trend);
  renderDailyChart(daily);
  loadPieChart('expense', document.querySelector('.tog.active'));
}
function renderSummaryCards(s) {
  const fmt = v => 'S/ ' + Math.abs(v).toLocaleString('es-PE',{minimumFractionDigits:2,maximumFractionDigits:2});
  const [y, m] = currentMonth.split('-');
  const months = ['Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre'];
  document.getElementById('summaryCards').innerHTML = `
    <div class="card income">
      <div class="card-label">Ingresos</div>
      <div class="card-amount">${fmt(s.income)}</div>
      <div class="card-sub">${months[+m-1]} ${y}</div>
      <div class="card-icon">💰</div>
    </div>
    <div class="card expense">
      <div class="card-label">Gastos</div>
      <div class="card-amount">${fmt(s.expense)}</div>
      <div class="card-sub">${months[+m-1]} ${y}</div>
      <div class="card-icon">💸</div>
    </div>
    <div class="card balance">
      <div class="card-label">Balance</div>
      <div class="card-amount" style="color:${s.balance>=0?'var(--income)':'var(--expense)'}">${s.balance>=0?'+':''}${fmt(s.balance)}</div>
      <div class="card-sub">Disponible</div>
      <div class="card-icon">⚖️</div>
    </div>`;
}
function renderTrendChart(data) {
  const ctx = document.getElementById('trendChart').getContext('2d');
  if (charts.trend) charts.trend.destroy();
  charts.trend = new Chart(ctx, {
    type:'line',
    data:{
      labels: data.months.map(m => { const [y,mo]=m.split('-'); return ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic'][+mo-1]+' '+y.slice(2); }),
      datasets:[
        {label:'Ingresos',data:data.income, borderColor:'#10b981',backgroundColor:'rgba(16,185,129,.08)',tension:.35,fill:true,pointRadius:4,pointHoverRadius:6},
        {label:'Gastos',  data:data.expense,borderColor:'#f43f5e',backgroundColor:'rgba(244,63,94,.08)', tension:.35,fill:true,pointRadius:4,pointHoverRadius:6},
      ]
    },
    options: chartOpts()
  });
}
async function loadPieChart(type, btn) {
  if (btn) {
    document.querySelectorAll('.tog').forEach(t=>t.classList.remove('active'));
    btn.classList.add('active');
  }
  const data = await fetch(`/api/stats/by_category?month=${currentMonth}&type=${type}`).then(r=>r.json());
  const ctx  = document.getElementById('pieChart').getContext('2d');
  if (charts.pie) charts.pie.destroy();
  const colors = ['#10b981','#3b82f6','#8b5cf6','#f59e0b','#ef4444','#ec4899','#14b8a6','#f97316','#6366f1','#a855f7','#22c55e','#0ea5e9'];
  charts.pie = new Chart(ctx, {
    type:'doughnut',
    data:{
      labels: data.map(d=>d.category),
      datasets:[{data:data.map(d=>d.total),backgroundColor:colors,hoverOffset:6,borderWidth:2,borderColor:'#111827'}]
    },
    options:{...pieOpts(),plugins:{legend:{position:'right',labels:{color:'#94a3b8',font:{size:11}}}}}
  });
}
function renderDailyChart(data) {
  const ctx = document.getElementById('dailyChart').getContext('2d');
  if (charts.daily) charts.daily.destroy();
  charts.daily = new Chart(ctx, {
    type:'bar',
    data:{
      labels: data.dates.map(d=>d.slice(5)),
      datasets:[
        {label:'Ingresos',data:data.income, backgroundColor:'rgba(16,185,129,.7)',borderRadius:4},
        {label:'Gastos',  data:data.expense,backgroundColor:'rgba(244,63,94,.7)', borderRadius:4},
      ]
    },
    options:{...chartOpts(),scales:{x:{...barXScale()},y:{...barYScale()}}}
  });
}

async function loadReports() {
  const [catExp, catInc, trend] = await Promise.all([
    fetch(`/api/stats/by_category?month=${currentMonth}&type=expense`).then(r=>r.json()),
    fetch(`/api/stats/by_category?month=${currentMonth}&type=income`).then(r=>r.json()),
    fetch('/api/stats/monthly_trend').then(r=>r.json()),
  ]);
  renderBarChart('barCatChart', catExp, 'barCat', 'rgba(244,63,94,.7)');
  renderBarChart('barIncChart', catInc, 'barInc', 'rgba(16,185,129,.7)');
  renderAccumChart(trend);
}
function renderBarChart(id, data, key, color) {
  const ctx = document.getElementById(id).getContext('2d');
  if (charts[key]) charts[key].destroy();
  charts[key] = new Chart(ctx, {
    type:'bar',
    data:{
      labels:data.map(d=>d.category),
      datasets:[{label:'S/',data:data.map(d=>d.total),backgroundColor:color,borderRadius:6}]
    },
    options:{...chartOpts(),indexAxis:'y',scales:{x:{...barYScale()},y:{...barXScale()}}}
  });
}
function renderAccumChart(data) {
  const ctx = document.getElementById('accumChart').getContext('2d');
  if (charts.accum) charts.accum.destroy();
  let accum = 0;
  const accumData = data.months.map((_,i) => { accum += data.income[i] - data.expense[i]; return +accum.toFixed(2); });
  charts.accum = new Chart(ctx, {
    type:'line',
    data:{
      labels:data.months.map(m => { const [y,mo]=m.split('-'); return ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic'][+mo-1]+' '+y.slice(2); }),
      datasets:[{label:'Balance acumulado',data:accumData,borderColor:'#3b82f6',backgroundColor:'rgba(59,130,246,.08)',tension:.35,fill:true,pointRadius:4}]
    },
    options: chartOpts()
  });
}

async function loadTransactions() {
  const params = new URLSearchParams();
  const search = document.getElementById('searchInput')?.value.trim();
  const type   = document.getElementById('filterType')?.value;
  const cat    = document.getElementById('filterCat')?.value;
  const from   = document.getElementById('filterFrom')?.value;
  const to     = document.getElementById('filterTo')?.value;
  if (search) params.set('search', search);
  if (type)   params.set('type',   type);
  if (cat)    params.set('category', cat);
  if (from)   params.set('date_from', from);
  if (to)     params.set('date_to', to);
  const txs = await fetch('/api/transactions?' + params).then(r=>r.json());
  const list = document.getElementById('txList');
  if (!txs.length) {
    list.innerHTML = '<div class="empty-state"><div class="big">📭</div><div>Sin movimientos encontrados</div></div>';
    return;
  }
  list.innerHTML = txs.map(tx => {
    const cat  = categories.find(c=>c.name===tx.category);
    const icon = cat?.icon || (tx.type==='income'?'💰':'💸');
    const fmt  = v => 'S/ ' + v.toLocaleString('es-PE',{minimumFractionDigits:2,maximumFractionDigits:2});
    return `
    <div class="tx-item">
      <div class="tx-icon ${tx.type}">${icon}</div>
      <div class="tx-info">
        <div class="tx-cat">${tx.category}</div>
        <div class="tx-desc">${tx.description || '—'}</div>
      </div>
      <div class="tx-date">${tx.date}</div>
      <div class="tx-amount ${tx.type}">${tx.type==='income'?'+':'-'}${fmt(tx.amount)}</div>
      <div class="tx-actions">
        <button class="tx-btn edit" onclick='openEditModal(${JSON.stringify(tx)})' title="Editar">✏</button>
        <button class="tx-btn del"  onclick="deleteTransaction(${tx.id})" title="Eliminar">🗑</button>
      </div>
    </div>`;
  }).join('');
}
async function deleteTransaction(id) {
  if (!confirm('¿Eliminar este movimiento?')) return;
  await fetch(`/api/transactions/${id}`, {method:'DELETE'});
  showToast('Movimiento eliminado');
  loadTransactions();
  if (document.getElementById('sec-dashboard').classList.contains('active')) loadDashboard();
}

function openAddModal() {
  document.getElementById('editId').value = '';
  document.getElementById('fAmount').value = '';
  document.getElementById('fDescription').value = '';
  document.getElementById('fDate').value = new Date().toISOString().slice(0,10);
  document.getElementById('modalTitle').textContent = 'Nueva transacción';
  document.getElementById('submitBtn').textContent  = 'Guardar';
  setType('expense', document.querySelector('[data-type="expense"]'));
  document.getElementById('modalBg').classList.add('active');
}
function openEditModal(tx) {
  document.getElementById('editId').value = tx.id;
  document.getElementById('fAmount').value = tx.amount;
  document.getElementById('fDescription').value = tx.description || '';
  document.getElementById('fDate').value = tx.date;
  document.getElementById('modalTitle').textContent = 'Editar movimiento';
  document.getElementById('submitBtn').textContent  = 'Actualizar';
  setType(tx.type, document.querySelector(`[data-type="${tx.type}"]`));
  setTimeout(() => { document.getElementById('fCategory').value = tx.category; }, 50);
  document.getElementById('modalBg').classList.add('active');
}
function closeModal() { document.getElementById('modalBg').classList.remove('active'); }
document.getElementById('modalBg').addEventListener('click', e => { if (e.target === e.currentTarget) closeModal(); });

function setType(type, btn) {
  currentType = type;
  document.querySelectorAll('.type-tab').forEach(t => t.classList.remove('active'));
  if (btn) btn.classList.add('active');
  populateCategorySelects();
}
async function submitTransaction() {
  const id     = document.getElementById('editId').value;
  const amount = parseFloat(document.getElementById('fAmount').value);
  const cat    = document.getElementById('fCategory').value;
  const date   = document.getElementById('fDate').value;
  const desc   = document.getElementById('fDescription').value.trim();
  if (!amount || amount <= 0) { showToast('Ingresa un monto válido','error'); return; }
  if (!cat)  { showToast('Selecciona una categoría','error'); return; }
  if (!date) { showToast('Selecciona una fecha','error'); return; }
  const body = { type:currentType, amount, category:cat, date, description:desc };
  const url    = id ? `/api/transactions/${id}` : '/api/transactions';
  const method = id ? 'PUT' : 'POST';
  const res = await fetch(url, { method, headers:{'Content-Type':'application/json'}, body:JSON.stringify(body) });
  if (res.ok) {
    closeModal();
    showToast(id ? 'Movimiento actualizado ✓' : 'Movimiento registrado ✓');
    const active = document.querySelector('.nav-btn.active')?.dataset.section;
    if (active === 'transactions') loadTransactions();
    if (active === 'dashboard')    loadDashboard();
    if (active === 'reports')      loadReports();
  } else {
    showToast('Error al guardar','error');
  }
}

function exportExcel() { showToast('Generando Excel…'); window.location.href = '/api/export/excel'; }

function showToast(msg, type='') {
  const t = document.getElementById('toast');
  t.textContent = msg;
  t.className   = 'toast show' + (type ? ' ' + type : '');
  setTimeout(() => t.classList.remove('show'), 2800);
}

const gridColor  = 'rgba(255,255,255,.05)';
const labelColor = '#64748b';
function chartOpts() {
  return {
    responsive:true,
    plugins:{
      legend:{labels:{color:labelColor,font:{size:11}}},
      tooltip:{backgroundColor:'#1e2740',titleColor:'#f1f5f9',bodyColor:'#94a3b8',
               callbacks:{label:ctx => ' S/ '+(ctx.parsed.y?.toLocaleString('es-PE',{minimumFractionDigits:2})||'')}}
    },
    scales:{
      x:{grid:{color:gridColor},ticks:{color:labelColor,font:{size:10}}},
      y:{grid:{color:gridColor},ticks:{color:labelColor,font:{size:10},callback:v=>'S/'+v.toLocaleString('es-PE',{maximumFractionDigits:0})}}
    }
  };
}
function pieOpts() {
  return {
    responsive:true,
    plugins:{
      legend:{labels:{color:labelColor,font:{size:11}}},
      tooltip:{backgroundColor:'#1e2740',titleColor:'#f1f5f9',bodyColor:'#94a3b8',
               callbacks:{label:ctx=>' S/ '+ctx.parsed.toLocaleString('es-PE',{minimumFractionDigits:2})}}
    }
  };
}
function barXScale() { return {grid:{color:gridColor},ticks:{color:labelColor,font:{size:10}}}; }
function barYScale() {
  return {grid:{color:gridColor},ticks:{color:labelColor,font:{size:10},callback:v=>'S/'+Number(v).toLocaleString('es-PE',{maximumFractionDigits:0})}};
}
</script>
</body>
</html>'''

AUTH_HTML = '''<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Finanzas – {title}</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Sans:wght@300;400;500&display=swap" rel="stylesheet">
<style>
:root{{--bg:#0a0e1a;--bg2:#111827;--bg3:#1e2740;--border:#1e293b;--border2:#334155;--text:#f1f5f9;--text2:#94a3b8;--text3:#64748b;--accent:#6ee7b7;--accent2:#34d399;--income:#10b981;--expense:#f43f5e;--radius:14px}}
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
body{{background:var(--bg);color:var(--text);font-family:'DM Sans',sans-serif;min-height:100vh;display:flex;align-items:center;justify-content:center;padding:20px}}
body::before{{content:'';position:fixed;inset:0;background:radial-gradient(ellipse at 20% 50%,rgba(110,231,183,.06) 0%,transparent 60%),radial-gradient(ellipse at 80% 20%,rgba(139,92,246,.06) 0%,transparent 50%);pointer-events:none}}
button{{cursor:pointer;border:none;font-family:inherit}}
input{{font-family:inherit}}
.box{{background:var(--bg2);border:1px solid var(--border2);border-radius:24px;padding:44px 40px;width:min(440px,100%);position:relative;animation:authIn .4s cubic-bezier(.34,1.56,.64,1)}}
@keyframes authIn{{from{{opacity:0;transform:translateY(20px) scale(.97)}}to{{opacity:1;transform:none}}}}
.logo{{display:flex;align-items:center;gap:10px;margin-bottom:32px;justify-content:center}}
.logo-icon{{font-size:28px;color:var(--accent)}}
.logo-text{{font-family:'Syne',sans-serif;font-weight:800;font-size:22px;letter-spacing:.03em}}
h1{{font-family:'Syne',sans-serif;font-size:24px;font-weight:800;margin-bottom:6px;text-align:center}}
.sub{{font-size:14px;color:var(--text3);text-align:center;margin-bottom:28px}}
.form{{display:flex;flex-direction:column;gap:14px}}
.group{{display:flex;flex-direction:column;gap:6px}}
.group label{{font-size:11px;font-weight:700;color:var(--text3);text-transform:uppercase;letter-spacing:.08em}}
.group input{{background:var(--bg3);border:1px solid var(--border2);color:var(--text);padding:12px 16px;border-radius:12px;font-size:14px;outline:none;transition:border-color .15s}}
.group input:focus{{border-color:var(--accent)}}
.btn{{background:linear-gradient(135deg,var(--accent),#059669);color:#064e3b;font-weight:800;font-size:15px;padding:14px;border-radius:12px;margin-top:8px;transition:all .2s;width:100%}}
.btn:hover{{filter:brightness(1.08);transform:translateY(-1px);box-shadow:0 6px 20px rgba(16,185,129,.3)}}
.link{{text-align:center;font-size:13px;color:var(--text3);margin-top:20px}}
.link a{{color:var(--accent);font-weight:600;text-decoration:none}}
.link a:hover{{text-decoration:underline}}
.err{{background:rgba(244,63,94,.12);border:1px solid rgba(244,63,94,.3);color:#f87171;padding:10px 14px;border-radius:10px;font-size:13px;text-align:center}}
.divider{{display:flex;align-items:center;gap:10px;margin:4px 0}}
.divider::before,.divider::after{{content:'';flex:1;height:1px;background:var(--border)}}
.divider span{{font-size:11px;color:var(--text3)}}
.features{{display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-bottom:28px}}
.feat{{background:var(--bg3);border:1px solid var(--border);border-radius:10px;padding:10px 12px;display:flex;align-items:center;gap:8px;font-size:12px;color:var(--text2)}}
.feat span{{font-size:16px}}
</style>
</head>
<body>
<div class="box">
  <div class="logo">
    <span class="logo-icon">◈</span>
    <span class="logo-text">Finanzas</span>
  </div>
  {content}
</div>
</body>
</html>'''

# ── RUTAS AUTH ───────────────────────────────────────────────────────────────
@app.route('/login', methods=['GET'])
def login_page():
    if 'user_id' in session:
        return redirect('/')
    error = request.args.get('error', '')
    err_html = f'<div class="err">{error}</div>' if error else ''
    content = f'''
      <h1>Sea bienvenido</h1>
      <p class="sub">Monitorea tus finanzas personales</p>
      {err_html}
      <div class="form">
        <div class="group">
          <label>Email o usuario</label>
          <input type="text" id="identifier" placeholder="tu@email.com" autocomplete="username">
        </div>
        <div class="group">
          <label>Contraseña</label>
          <input type="password" id="password" placeholder="••••••••" autocomplete="current-password" onkeydown="if(event.key==='Enter')doLogin()">
        </div>
        <button class="btn" onclick="doLogin()">Iniciar sesión</button>
      </div>
      <div class="divider"><span>¿No tienes cuenta?</span></div>
      <p class="link"><a href="/register">Crear cuenta gratis →</a></p>
      <script>
      async function doLogin() {{
        const identifier = document.getElementById('identifier').value.trim();
        const password   = document.getElementById('password').value;
        if (!identifier || !password) {{ alert('Completa todos los campos'); return; }}
        const res = await fetch('/api/login', {{
          method:'POST', headers:{{'Content-Type':'application/json'}},
          body: JSON.stringify({{identifier, password}})
        }});
        if (res.ok) {{ window.location.href = '/'; }}
        else {{
          const d = await res.json();
          window.location.href = '/login?error=' + encodeURIComponent(d.error || 'Error al iniciar sesión');
        }}
      }}
      </script>'''
    return AUTH_HTML.format(title='Iniciar sesión', content=content)

@app.route('/register', methods=['GET'])
def register_page():
    if 'user_id' in session:
        return redirect('/')
    error = request.args.get('error', '')
    err_html = f'<div class="err">{error}</div>' if error else ''
    content = f'''
      <h1>Crear cuenta</h1>
      <p class="sub">Tu dinero, solo tuyo. Privado y seguro.</p>
      <div class="features">
        <div class="feat"><span>📊</span> Dashboard en tiempo real</div>
        <div class="feat"><span>🏷️</span> Categorías personalizadas</div>
        <div class="feat"><span>📈</span> Reportes mensuales</div>
        <div class="feat"><span>📥</span> Exportar a Excel</div>
      </div>
      {err_html}
      <div class="form">
        <div class="group">
          <label>Nombre de usuario</label>
          <input type="text" id="username" placeholder="juanperez" autocomplete="username">
        </div>
        <div class="group">
          <label>Email</label>
          <input type="email" id="email" placeholder="tu@email.com" autocomplete="email">
        </div>
        <div class="group">
          <label>Contraseña</label>
          <input type="password" id="password" placeholder="Mínimo 6 caracteres" autocomplete="new-password">
        </div>
        <div class="group">
          <label>Confirmar contraseña</label>
          <input type="password" id="password2" placeholder="Repite tu contraseña" autocomplete="new-password" onkeydown="if(event.key==='Enter')doRegister()">
        </div>
        <button class="btn" onclick="doRegister()">Crear mi cuenta</button>
      </div>
      <div class="divider"><span>¿Ya tienes cuenta?</span></div>
      <p class="link"><a href="/login">Iniciar sesión →</a></p>
      <script>
      async function doRegister() {{
        const username  = document.getElementById('username').value.trim();
        const email     = document.getElementById('email').value.trim();
        const password  = document.getElementById('password').value;
        const password2 = document.getElementById('password2').value;
        if (!username || !email || !password) {{ alert('Completa todos los campos'); return; }}
        if (password !== password2) {{ window.location.href='/register?error=Las contraseñas no coinciden'; return; }}
        if (password.length < 6) {{ window.location.href='/register?error=La contraseña debe tener al menos 6 caracteres'; return; }}
        const res = await fetch('/api/register', {{
          method:'POST', headers:{{'Content-Type':'application/json'}},
          body: JSON.stringify({{username, email, password}})
        }});
        if (res.ok) {{ window.location.href = '/'; }}
        else {{
          const d = await res.json();
          window.location.href = '/register?error=' + encodeURIComponent(d.error || 'Error al registrar');
        }}
      }}
      </script>'''
    return AUTH_HTML.format(title='Crear cuenta', content=content)

# ── API AUTH ─────────────────────────────────────────────────────────────────
@app.route('/api/register', methods=['POST'])
def api_register():
    d = request.get_json()
    username = d.get('username', '').strip()
    email    = d.get('email', '').strip().lower()
    password = d.get('password', '')
    if not username or not email or not password:
        return jsonify({'error': 'Todos los campos son requeridos'}), 400
    if len(password) < 6:
        return jsonify({'error': 'La contraseña debe tener al menos 6 caracteres'}), 400
    conn = get_db()
    try:
        cur = conn.execute(
            'INSERT INTO users (username, email, password) VALUES (?,?,?)',
            (username, email, hash_password(password))
        )
        user_id = cur.lastrowid
        conn.commit()
        seed_user_categories(user_id)
        session['user_id']  = user_id
        session['username'] = username
        conn.close()
        return jsonify({'ok': True, 'username': username}), 201
    except sqlite3.IntegrityError as e:
        conn.close()
        if 'username' in str(e):
            return jsonify({'error': 'El nombre de usuario ya está en uso'}), 400
        return jsonify({'error': 'El email ya está registrado'}), 400

@app.route('/api/login', methods=['POST'])
def api_login():
    d = request.get_json()
    identifier = d.get('identifier', '').strip().lower()
    password   = d.get('password', '')
    conn = get_db()
    row = conn.execute(
        'SELECT * FROM users WHERE (LOWER(email)=? OR LOWER(username)=?) AND password=?',
        (identifier, identifier, hash_password(password))
    ).fetchone()
    conn.close()
    if not row:
        return jsonify({'error': 'Credenciales incorrectas'}), 401
    session['user_id']  = row['id']
    session['username'] = row['username']
    return jsonify({'ok': True, 'username': row['username']})

@app.route('/api/logout', methods=['POST'])
def api_logout():
    session.clear()
    return jsonify({'ok': True})

@app.route('/api/me')
@login_required
def api_me():
    return jsonify({'user_id': current_user_id(), 'username': session.get('username')})

# ── MAIN PAGE ────────────────────────────────────────────────────────────────
@app.route('/')
@login_required
def index():
    return HTML

# ── API TRANSACCIONES ────────────────────────────────────────────────────────
@app.route('/api/transactions', methods=['GET'])
@login_required
def get_transactions():
    uid = current_user_id()
    conn = get_db()
    filters, params = ['user_id = ?'], [uid]
    if request.args.get('type'):
        filters.append('type = ?'); params.append(request.args['type'])
    if request.args.get('category'):
        filters.append('category = ?'); params.append(request.args['category'])
    if request.args.get('date_from'):
        filters.append('date >= ?'); params.append(request.args['date_from'])
    if request.args.get('date_to'):
        filters.append('date <= ?'); params.append(request.args['date_to'])
    if request.args.get('search'):
        filters.append('(description LIKE ? OR category LIKE ?)')
        params += [f'%{request.args["search"]}%'] * 2
    where = 'WHERE ' + ' AND '.join(filters)
    rows = conn.execute(f'SELECT * FROM transactions {where} ORDER BY date DESC, id DESC', params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/transactions', methods=['POST'])
@login_required
def add_transaction():
    uid = current_user_id()
    d = request.get_json()
    conn = get_db()
    cur = conn.execute(
        'INSERT INTO transactions (user_id,type,amount,category,description,date) VALUES (?,?,?,?,?,?)',
        (uid, d['type'], float(d['amount']), d['category'], d.get('description',''), d['date'])
    )
    conn.commit()
    row = conn.execute('SELECT * FROM transactions WHERE id=?', (cur.lastrowid,)).fetchone()
    conn.close()
    return jsonify(dict(row)), 201

@app.route('/api/transactions/<int:tid>', methods=['PUT'])
@login_required
def update_transaction(tid):
    uid = current_user_id()
    d = request.get_json()
    conn = get_db()
    conn.execute(
        'UPDATE transactions SET type=?,amount=?,category=?,description=?,date=? WHERE id=? AND user_id=?',
        (d['type'], float(d['amount']), d['category'], d.get('description',''), d['date'], tid, uid)
    )
    conn.commit()
    row = conn.execute('SELECT * FROM transactions WHERE id=? AND user_id=?', (tid, uid)).fetchone()
    conn.close()
    return jsonify(dict(row))

@app.route('/api/transactions/<int:tid>', methods=['DELETE'])
@login_required
def delete_transaction(tid):
    uid = current_user_id()
    conn = get_db()
    conn.execute('DELETE FROM transactions WHERE id=? AND user_id=?', (tid, uid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ── API CATEGORÍAS ───────────────────────────────────────────────────────────
@app.route('/api/categories', methods=['GET'])
@login_required
def get_categories():
    uid = current_user_id()
    conn = get_db()
    rows = conn.execute('SELECT * FROM categories WHERE user_id=? ORDER BY type, name', (uid,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/categories', methods=['POST'])
@login_required
def add_category():
    uid = current_user_id()
    d = request.get_json()
    conn = get_db()
    try:
        cur = conn.execute('INSERT INTO categories (user_id,name,type,icon,color) VALUES (?,?,?,?,?)',
                           (uid, d['name'], d['type'], d.get('icon','📁'), d.get('color','#6366f1')))
        conn.commit()
        row = conn.execute('SELECT * FROM categories WHERE id=?', (cur.lastrowid,)).fetchone()
        conn.close()
        return jsonify(dict(row)), 201
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'error': 'Categoría ya existe'}), 400

@app.route('/api/categories/<int:cid>', methods=['DELETE'])
@login_required
def delete_category(cid):
    uid = current_user_id()
    conn = get_db()
    conn.execute('DELETE FROM categories WHERE id=? AND user_id=?', (cid, uid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ── API ESTADÍSTICAS ─────────────────────────────────────────────────────────
@app.route('/api/stats/summary')
@login_required
def stats_summary():
    uid = current_user_id()
    conn = get_db()
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    rows = conn.execute(
        "SELECT type, SUM(amount) as total FROM transactions WHERE user_id=? AND strftime('%Y-%m',date)=? GROUP BY type",
        (uid, month)).fetchall()
    conn.close()
    data = {r['type']: r['total'] for r in rows}
    income  = data.get('income', 0)
    expense = data.get('expense', 0)
    return jsonify({'income': income, 'expense': expense, 'balance': income - expense, 'month': month})

@app.route('/api/stats/by_category')
@login_required
def stats_by_category():
    uid = current_user_id()
    conn = get_db()
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    ttype = request.args.get('type', 'expense')
    rows = conn.execute(
        "SELECT category, SUM(amount) as total FROM transactions "
        "WHERE user_id=? AND type=? AND strftime('%Y-%m',date)=? GROUP BY category ORDER BY total DESC",
        (uid, ttype, month)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/stats/monthly_trend')
@login_required
def monthly_trend():
    uid = current_user_id()
    conn = get_db()
    rows = conn.execute(
        "SELECT strftime('%Y-%m',date) as month, type, SUM(amount) as total "
        "FROM transactions WHERE user_id=? GROUP BY month, type ORDER BY month",
        (uid,)).fetchall()
    conn.close()
    months = sorted(set(r['month'] for r in rows))[-12:]
    income  = {m: 0 for m in months}
    expense = {m: 0 for m in months}
    for r in rows:
        if r['month'] in months:
            if r['type'] == 'income':  income[r['month']]  = r['total']
            else:                       expense[r['month']] = r['total']
    return jsonify({'months': months, 'income': [income[m] for m in months], 'expense': [expense[m] for m in months]})

@app.route('/api/stats/daily_trend')
@login_required
def daily_trend():
    uid = current_user_id()
    conn = get_db()
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    rows = conn.execute(
        "SELECT date, type, SUM(amount) as total FROM transactions "
        "WHERE user_id=? AND strftime('%Y-%m',date)=? GROUP BY date, type ORDER BY date",
        (uid, month)).fetchall()
    conn.close()
    data = defaultdict(lambda: {'income': 0, 'expense': 0})
    for r in rows:
        data[r['date']][r['type']] = r['total']
    dates = sorted(data.keys())
    return jsonify({'dates': dates, 'income': [data[d]['income'] for d in dates], 'expense': [data[d]['expense'] for d in dates]})

# ── EXPORTAR EXCEL ───────────────────────────────────────────────────────────
@app.route('/api/export/excel')
@login_required
def export_excel():
    uid = current_user_id()
    conn = get_db()
    rows = conn.execute('SELECT * FROM transactions WHERE user_id=? ORDER BY date DESC', (uid,)).fetchall()
    cats = conn.execute('SELECT * FROM categories WHERE user_id=?', (uid,)).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Transacciones'
    header_fill = PatternFill('solid', fgColor='1E293B')
    income_fill  = PatternFill('solid', fgColor='D1FAE5')
    expense_fill = PatternFill('solid', fgColor='FEE2E2')
    header_font  = Font(bold=True, color='FFFFFF', size=11)
    thin = Border(
        left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),  bottom=Side(style='thin', color='CBD5E1'),
    )
    headers = ['#','Tipo','Fecha','Categoría','Descripción','Monto (S/)']
    col_widths = [5, 12, 14, 20, 35, 15]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

    total_income = total_expense = 0
    for i, r in enumerate(rows, 2):
        fill = income_fill if r['type'] == 'income' else expense_fill
        vals = [r['id'], 'Ingreso' if r['type']=='income' else 'Gasto',
                r['date'], r['category'], r['description'] or '', r['amount']]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.fill = fill; cell.border = thin
            cell.alignment = Alignment(vertical='center')
            if col == 6:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
        if r['type'] == 'income': total_income  += r['amount']
        else:                      total_expense += r['amount']

    last = len(rows) + 2
    ws.cell(row=last,   column=5, value='TOTAL INGRESOS').font = Font(bold=True, color='065F46')
    ws.cell(row=last,   column=6, value=total_income).number_format = '#,##0.00'
    ws.cell(row=last+1, column=5, value='TOTAL GASTOS').font = Font(bold=True, color='991B1B')
    ws.cell(row=last+1, column=6, value=total_expense).number_format = '#,##0.00'
    ws.cell(row=last+2, column=5, value='BALANCE').font = Font(bold=True)
    bal = ws.cell(row=last+2, column=6, value=total_income - total_expense)
    bal.number_format = '#,##0.00'; bal.font = Font(bold=True, color='1D4ED8')
    ws.freeze_panes = 'A2'

    ws2 = wb.create_sheet('Resumen Mensual')
    conn2 = get_db()
    monthly = conn2.execute(
        "SELECT strftime('%Y-%m',date) as month, type, SUM(amount) as total "
        "FROM transactions WHERE user_id=? GROUP BY month,type ORDER BY month", (uid,)
    ).fetchall()
    conn2.close()
    months_data = defaultdict(lambda: {'income':0,'expense':0})
    for r in monthly:
        months_data[r['month']][r['type']] = r['total']
    for col, h in enumerate(['Mes','Ingresos (S/)','Gastos (S/)','Balance (S/)'], 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center'); cell.border = thin
        ws2.column_dimensions[get_column_letter(col)].width = 18
    for i, m in enumerate(sorted(months_data.keys()), 2):
        inc = months_data[m]['income']; exp = months_data[m]['expense']
        for col, v in enumerate([m, inc, exp, inc-exp], 1):
            cell = ws2.cell(row=i, column=col, value=v)
            cell.border = thin
            if col > 1: cell.number_format = '#,##0.00'

    ws3 = wb.create_sheet('Categorías')
    for col, h in enumerate(['ID','Nombre','Tipo','Ícono'], 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws3.column_dimensions[get_column_letter(col)].width = 20
    for i, c in enumerate(cats, 2):
        ws3.cell(row=i,column=1,value=c['id']); ws3.cell(row=i,column=2,value=c['name'])
        ws3.cell(row=i,column=3,value='Ingreso' if c['type']=='income' else 'Gasto' if c['type']=='expense' else 'Ambos')
        ws3.cell(row=i,column=4,value=c['icon'])

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"gastos_{session.get('username','user')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)

if __name__ == '__main__':
    init_db()
    app.run(debug=True, port=5000)
